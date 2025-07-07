#!/usr/bin/env python3
"""Excel Analyzer Module"""
import pandas as pd
import openpyxl

class ExcelAnalyzer:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        return self.workbook

    def read_sheet(self, sheet_name):
        return pd.read_excel(self.file_path, sheet_name=sheet_name)

    def analyze_sheet(self, sheet_name):
        df = self.read_sheet(sheet_name)
        analysis = {
            'row_count': len(df),
            'column_count': len(df.columns),
            'missing_values': df.isnull().sum().sum(),
            'numeric_columns': df.select_dtypes(include=['number']).columns.tolist(),
            'text_columns': df.select_dtypes(include=['object']).columns.tolist(),
            'date_columns': df.select_dtypes(include=['datetime']).columns.tolist()
        }
        return analysis

    def get_sheet_names(self):
        if not self.workbook:
            self.load_workbook()
        return self.workbook.sheetnames

    def get_named_ranges(self):
        if not self.workbook:
            self.load_workbook()
        return self.workbook.defined_names.keys()

    def get_pivot_tables(self, sheet_name):
        # This is a simplified implementation
        # In a real scenario, you would need to parse the Excel file structure
        # to identify pivot tables
        return []

    def get_formulas(self, sheet_name):
        if not self.workbook:
            self.load_workbook()
        
        sheet = self.workbook[sheet_name]
        formulas = {}
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formulas[cell.coordinate] = cell.value
        
        return formulas

    def get_charts(self, sheet_name):
        if not self.workbook:
            self.load_workbook()
        
        sheet = self.workbook[sheet_name]
        return len(sheet._charts)
